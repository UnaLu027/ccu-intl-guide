#!/usr/bin/env python3
"""
compare.py — CCU campusData vs Excel 核查表 比對工具

用法：
    python3 compare.py

需求套件：
    pip install openpyxl

放置位置：專案根目錄（與 client/ 同層）
輸入：
    - campusData_核查表_1_.xlsx（放在專案根目錄）
    - client/src/data/campusData.ts
輸出：
    - 終端機差異清單 + campusData 修改片段
    - compare_result.xlsx（差異標紅）
"""

import json, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH      = Path("campusData_核查表_1_.xlsx")
CAMPUSDATA_PATH = Path("client/src/data/campusData.ts")
OUTPUT_PATH     = Path("compare_result.xlsx")

# ── Styles ──────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FFCCCC")
YELLOW_FILL = PatternFill("solid", start_color="FFF2CC")
GREEN_FILL  = PatternFill("solid", start_color="D9F0D3")
NAVY_FILL   = PatternFill("solid", start_color="1B2A4A")
SAGE_FILL   = PatternFill("solid", start_color="7A9E7E")
TEAL_FILL   = PatternFill("solid", start_color="4A8E8E")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
GRAY_FILL   = PatternFill("solid", start_color="F5F5F5")

def wfont(color="FFFFFF", bold=True, size=9):
    return Font(color=color, bold=bold, name="Arial", size=size)

def border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(h="left"):
    return Alignment(horizontal=h, vertical="top", wrap_text=True)

# ── Field mappings ───────────────────────────────────────────
OFFICE_FIELDS = {
    "中文名稱":             "name_zh",
    "英文名稱":             "name_en",
    "樓層":                 "floor",
    "室號（中）":           "room_zh",
    "室號（英）":           "room_en",
    "完整地點顯示（中）★":  "indoor_location_note_zh",
    "完整地點顯示（英）★":  "indoor_location_note_en",
    "大樓（中）":           "building_name_zh",
    "大樓（英）":           "building_name_en",
    "功能描述（中）★":      "function_desc_zh",
    "功能描述（英）★":      "function_desc_en",
    "服務範圍（中）★":      "service_scope_zh",
    "服務範圍（英）★":      "service_scope_en",
    "常見情境（中）★":      "common_scenarios_zh",
    "常見情境（英）★":      "common_scenarios_en",
    "辦公時間":             "office_hours",
    "電話":                 "phone",
    "Email":                "email",
    "官網":                 "official_url",
}

DEPT_FIELDS = {
    "中文名稱":             "name_zh",
    "英文名稱":             "name_en",
    "學院（中）":           "college_zh",
    "學院（英）":           "college_en",
    "樓層":                 "floor",
    "室號（中）":           "room_zh",
    "室號（英）":           "room_en",
    "完整地點顯示（中）★":  "indoor_location_note_zh",
    "完整地點顯示（英）★":  "indoor_location_note_en",
    "大樓（中）":           "building_name_zh",
    "大樓（英）":           "building_name_en",
    "功能描述（中）★":      "function_desc_zh",
    "功能描述（英）★":      "function_desc_en",
    "服務範圍（中）★":      "service_scope_zh",
    "服務範圍（英）★":      "service_scope_en",
    "官網":                 "official_url",
}

TASK_FIELDS = {
    "任務名稱（中）★":      "task_name_zh",
    "任務名稱（英）★":      "task_name_en",
    "情境描述（中）★":      "scenario_zh",
    "情境描述（英）★":      "scenario_en",
}

# ── Helpers ──────────────────────────────────────────────────
def norm(v):
    if v is None: return ""
    s = str(v).strip()
    if s.lower() == "null": return ""
    return s

def parse_arr(content, marker):
    s = content.find(marker) + len(marker)
    depth=0; in_str=False; escape=False
    for i, ch in enumerate(content[s:], s):
        if escape: escape=False; continue
        if ch=='\\' and in_str: escape=True; continue
        if ch=='"' and not escape: in_str = not in_str; continue
        if in_str: continue
        if ch=='[': depth+=1
        elif ch==']':
            depth-=1
            if depth==0: return json.loads(content[s:i+1])
    return []

# ── Load data ────────────────────────────────────────────────
def load_campusdata():
    if not CAMPUSDATA_PATH.exists():
        print(f"❌ 找不到 {CAMPUSDATA_PATH}，請在專案根目錄執行"); sys.exit(1)
    c = CAMPUSDATA_PATH.read_text(encoding="utf-8")
    offices = parse_arr(c, "export const offices: Office[] = ")
    depts   = parse_arr(c, "export const departments: Department[] = ")
    tasks   = parse_arr(c, "export const tasks: Task[] = ")
    return (
        {o["id"]: o for o in offices},
        {d["id"]: d for d in depts},
        {t["id"]: t for t in tasks},
    )

def load_excel():
    if not EXCEL_PATH.exists():
        print(f"❌ 找不到 {EXCEL_PATH}"); sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH)

    def read_sheet(name):
        ws = wb[name]
        hdrs = [c.value for c in ws[2]]
        rows = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and row[1]:
                rows[str(row[0]).strip()] = {hdrs[i]: row[i] for i in range(min(len(hdrs), len(row)))}
        return rows

    offices_xl = {}
    ws1 = wb["行政處室"]
    h1 = [c.value for c in ws1[2]]
    for row in ws1.iter_rows(min_row=3, values_only=True):
        if row[0]:
            offices_xl[str(row[0]).strip()] = {h1[i]: row[i] for i in range(min(len(h1), len(row)))}

    depts_xl  = read_sheet("系所院辦")
    tasks_xl  = read_sheet("常見任務")
    return offices_xl, depts_xl, tasks_xl

# ── Compare ──────────────────────────────────────────────────
def compare(excel_rows, cd_rows, field_map):
    diffs, missing_in_cd, missing_in_xl = [], [], []
    for eid, xl_row in excel_rows.items():
        if eid not in cd_rows:
            missing_in_cd.append(eid); continue
        cd = cd_rows[eid]
        row_diffs = []
        for xl_col, cd_field in field_map.items():
            xv = norm(xl_row.get(xl_col))
            cv = norm(cd.get(cd_field))
            if xv == "" and cv == "": continue
            if xv != cv:
                row_diffs.append({"field": cd_field, "xl_col": xl_col, "xl": xv, "cd": cv})
        if row_diffs:
            name = norm(xl_row.get("中文名稱") or xl_row.get("任務名稱（中）★") or "")
            diffs.append({"id": eid, "name": name, "diffs": row_diffs})
    for cid in cd_rows:
        if cid not in excel_rows:
            missing_in_xl.append(cid)
    return diffs, missing_in_cd, missing_in_xl

# ── Terminal output ──────────────────────────────────────────
R="\033[91m"; Y="\033[93m"; G="\033[92m"; B="\033[1m"; E="\033[0m"

def print_results(diffs, miss_cd, miss_xl, cat):
    print(f"\n{'='*60}\n{B}{cat}{E}\n{'='*60}")
    if miss_cd:
        print(f"\n{R}⚠ Excel有 / campusData無（需新增）：{E}")
        for m in miss_cd: print(f"   - {m}")
    if miss_xl:
        print(f"\n{Y}⚠ campusData有 / Excel無（可考慮刪除）：{E}")
        for m in miss_xl: print(f"   - {m}")
    if not diffs:
        print(f"\n{G}✅ 所有欄位一致！{E}"); return
    print(f"\n{R}❌ {len(diffs)} 筆有差異：{E}\n")
    for item in diffs:
        print(f"  {B}[{item['id']}] {item['name']}{E}")
        for d in item["diffs"]:
            print(f"    {d['field']} ({d['xl_col']})")
            print(f"      Excel:      「{d['xl']}」")
            print(f"      campusData: 「{d['cd']}」")
        print()

def print_snippets(diffs, cat):
    if not diffs: return
    print(f"\n{'─'*60}\n修改片段 — {cat}\n{'─'*60}")
    for item in diffs:
        print(f"\n# {item['id']}")
        for d in item["diffs"]:
            val = d["xl"].replace('"', '\\"')
            print(f'  "{d["field"]}": "{val}",')

# ── Excel report ─────────────────────────────────────────────
def write_report(
    o_diffs, o_mcd, o_mxl,
    d_diffs, d_mcd, d_mxl,
    t_diffs, t_mcd, t_mxl,
    o_xl, d_xl, t_xl,
    o_cd, d_cd, t_cd,
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ────────────────────────────────────────
    ws0 = wb.create_sheet("摘要", 0)
    ws0.column_dimensions["A"].width = 35
    ws0.column_dimensions["B"].width = 10
    ws0.column_dimensions["C"].width = 38

    rows = [
        ("比對項目", "數量", "說明"),
        ("行政處室：欄位有差異", len(o_diffs), "見「行政處室」分頁"),
        ("行政處室：Excel有/campusData無", len(o_mcd), "需新增"),
        ("行政處室：campusData有/Excel無", len(o_mxl), "可考慮刪除"),
        ("系所院辦：欄位有差異", len(d_diffs), "見「系所院辦」分頁"),
        ("系所院辦：Excel有/campusData無", len(d_mcd), "需新增"),
        ("系所院辦：campusData有/Excel無", len(d_mxl), "可考慮刪除"),
        ("常見任務：欄位有差異", len(t_diffs), "見「常見任務」分頁"),
        ("常見任務：Excel有/campusData無", len(t_mcd), "需新增"),
        ("常見任務：campusData有/Excel無", len(t_mxl), "→ 應刪除（不在查核表內）"),
    ]
    for ri, (label, count, note) in enumerate(rows, 1):
        ws0.cell(ri, 1).value = label
        ws0.cell(ri, 2).value = count
        ws0.cell(ri, 3).value = note
        if ri == 1:
            for ci in range(1,4):
                ws0.cell(ri,ci).fill = NAVY_FILL
                ws0.cell(ri,ci).font = wfont()
                ws0.cell(ri,ci).alignment = Alignment(horizontal="center")
        else:
            is_bad = isinstance(count, int) and count > 0
            fill = RED_FILL if is_bad else GREEN_FILL
            for ci in range(1,4):
                ws0.cell(ri,ci).fill = fill
                ws0.cell(ri,ci).font = wfont("1B2A4A", False)
            ws0.cell(ri,2).alignment = Alignment(horizontal="center")
        for ci in range(1,4):
            ws0.cell(ri,ci).border = border()
        ws0.row_dimensions[ri].height = 18
    ws0.row_dimensions[1].height = 22

    # ── Data sheets ──────────────────────────────────────────
    def make_sheet(ws, label, diffs, miss_cd, miss_xl, xl_rows, cd_rows, field_map, hdr_fill):
        fkeys = list(field_map.keys())
        fvals = list(field_map.values())
        cols = ["id", "名稱", "狀態"] + [f"{k}\n(Excel)" for k in fkeys] + [f"{k}\n(campusData)" for k in fkeys]
        total_cols = len(cols)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(1,1).value = f"{label} 比對報告"
        ws.cell(1,1).fill = NAVY_FILL
        ws.cell(1,1).font = wfont(size=11)
        ws.cell(1,1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Header row
        ws.append(cols)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(2, ci)
            c.fill = hdr_fill
            c.font = wfont(size=8)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border()
        ws.row_dimensions[2].height = 36

        diff_map = {d["id"]: {x["field"]: x for x in d["diffs"]} for d in diffs}

        rn = 3
        def write_row(eid, name, status, xl_vals, cd_vals, row_fill, diff_fields=None):
            nonlocal rn
            row = [eid, name, status] + xl_vals + cd_vals
            ws.append(row)
            for ci in range(1, len(row)+1):
                c = ws.cell(rn, ci)
                c.fill = row_fill
                c.font = wfont("1B2A4A", False)
                c.border = border()
                c.alignment = wrap_align()
            if diff_fields:
                for fi, fv in enumerate(fvals):
                    if fv in diff_fields:
                        for ci in [3+fi+1, 3+len(fkeys)+fi+1]:
                            ws.cell(rn, ci).fill = RED_FILL
                            ws.cell(rn, ci).font = wfont("CC0000", True)
            ws.row_dimensions[rn].height = 40
            rn += 1

        for mid in miss_cd:
            xr = xl_rows.get(mid, {})
            xv = [norm(xr.get(k)) for k in fkeys]
            write_row(mid, norm(xr.get("中文名稱") or xr.get("任務名稱（中）★","")),
                      "❌ Excel有/campusData無", xv, ["—"]*len(fkeys), RED_FILL)

        for mid in miss_xl:
            cr = cd_rows.get(mid, {})
            cv = [norm(cr.get(f)) for f in fvals]
            write_row(mid, norm(cr.get("name_zh") or cr.get("task_name_zh","")),
                      "⚠ campusData有/Excel無", ["—"]*len(fkeys), cv, YELLOW_FILL)

        all_ids = sorted(set(list(xl_rows.keys()) + list(cd_rows.keys())))
        for eid in all_ids:
            if eid in miss_cd or eid in miss_xl: continue
            xr = xl_rows.get(eid, {})
            cr = cd_rows.get(eid, {})
            has_diff = eid in diff_map
            xv = [norm(xr.get(k)) for k in fkeys]
            cv = [norm(cr.get(f)) for f in fvals]
            status = "❌ 有差異" if has_diff else "✅ 一致"
            fill = PatternFill("solid", start_color="FFF0F0") if has_diff else (GRAY_FILL if rn%2==0 else WHITE_FILL)
            write_row(eid,
                      norm(xr.get("中文名稱") or xr.get("任務名稱（中）★") or cr.get("name_zh") or cr.get("task_name_zh","")),
                      status, xv, cv, fill,
                      diff_fields=set(diff_map[eid].keys()) if has_diff else None)

        # Column widths
        widths = [18, 20, 18] + [24]*len(fkeys) + [24]*len(fkeys)
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "D3"

    ws1 = wb.create_sheet("行政處室")
    make_sheet(ws1, "行政處室", o_diffs, o_mcd, o_mxl, o_xl, o_cd, OFFICE_FIELDS, SAGE_FILL)

    ws2 = wb.create_sheet("系所院辦")
    make_sheet(ws2, "系所院辦", d_diffs, d_mcd, d_mxl, d_xl, d_cd, DEPT_FIELDS, SAGE_FILL)

    ws3 = wb.create_sheet("常見任務")
    make_sheet(ws3, "常見任務", t_diffs, t_mcd, t_mxl, t_xl, t_cd, TASK_FIELDS, TEAL_FILL)

    wb.save(OUTPUT_PATH)
    print(f"\n📊 報告已輸出：{OUTPUT_PATH}")

# ── Main ─────────────────────────────────────────────────────
def main():
    print("🔍 載入 campusData.ts...")
    o_cd, d_cd, t_cd = load_campusdata()
    print(f"   Offices:{len(o_cd)}  Depts:{len(d_cd)}  Tasks:{len(t_cd)}")

    print("📋 載入 Excel 核查表...")
    o_xl, d_xl, t_xl = load_excel()
    print(f"   行政處室:{len(o_xl)}  系所院辦:{len(d_xl)}  常見任務:{len(t_xl)}")

    print("\n⚙️  比對中...")
    o_diffs, o_mcd, o_mxl = compare(o_xl, o_cd, OFFICE_FIELDS)
    d_diffs, d_mcd, d_mxl = compare(d_xl, d_cd, DEPT_FIELDS)
    t_diffs, t_mcd, t_mxl = compare(t_xl, t_cd, TASK_FIELDS)

    print_results(o_diffs, o_mcd, o_mxl, "行政處室")
    print_results(d_diffs, d_mcd, d_mxl, "系所院辦")
    print_results(t_diffs, t_mcd, t_mxl, "常見任務")

    print_snippets(o_diffs, "行政處室")
    print_snippets(d_diffs, "系所院辦")
    print_snippets(t_diffs, "常見任務")

    write_report(
        o_diffs, o_mcd, o_mxl,
        d_diffs, d_mcd, d_mxl,
        t_diffs, t_mcd, t_mxl,
        o_xl, d_xl, t_xl,
        o_cd, d_cd, t_cd,
    )

    total = len(o_diffs)+len(d_diffs)+len(t_diffs)
    missing = len(o_mcd)+len(d_mcd)+len(t_mcd)
    extra   = len(o_mxl)+len(d_mxl)+len(t_mxl)
    print(f"\n{'='*60}")
    print(f"差異：{total} 筆  需新增：{missing} 筆  多餘（可刪）：{extra} 筆")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
